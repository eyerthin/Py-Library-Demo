import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# 从控制台中获取路径和excel文件
path = input("请输入指定的路径：")
xlsx = input("请输入指定的xlsx文件：")

# 遍历文件夹中的所有文件的文件名，保存在列表file_name_list中
file_name_list = os.listdir(path=path)

# 将文件名的后缀去掉，并将新列表保存为file_n
file_n = list()
for i in range(len(file_name_list)):
    file_n.append(file_name_list[i].split('.')[0])

# 打开excel文件
wb = load_workbook(xlsx)
sheet_ranges = wb['Sheet1']
# 设置黄色背景色填充
fill = PatternFill('solid', fgColor="FFFF00")

# 分别为已标记的计数，未标记的计数，已标记的列表，未标记的文件名存储的列表。
count = 0
no_count = 0
reco = list()
no_reco = list()

# 遍历file_n中的文件名，并在sheet1第三列中寻找是否有单元格相等。
for name in file_n:
    for row in sheet_ranges.iter_cols(min_col=3,max_col=3):
        for cell in row:
            if name == cell.value:
                # 在sheet1第三列中找到对应文件名的单元格的话，将文件名append进reco列表中。
                reco.append(name)
                # 填充对应单元格
                c = 'C' + str(cell.row)
                sheet_ranges[c].fill = fill
                #计数
                count = count + 1
                break
        else:
            # 遍历sheet1的第三列结束，未找到相同单元格，记录进未记录。
            no_reco.append(name)
            no_count = no_count + 1

# 将结果回显到控制台
print('已标记{}'.format(reco))
print('已标记共{}项'.format(count))
print('未标记{}项'.format(no_count))
print("其中：{}，未在表格中找到".format(no_reco))

# 将标记结束的excel文件保存
wb.save('verification.xlsx')